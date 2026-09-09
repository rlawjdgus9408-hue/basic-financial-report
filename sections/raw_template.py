"""
기존 RAW 엑셀에 새 연도를 이어붙이거나, 처음부터 새 RAW 엑셀을 만든다.
"""
import io
import re
from copy import copy

import openpyxl
from openpyxl.utils import get_column_letter

_SECTION_BS = "재무상태표"
_SECTION_IS = "손익계산서"
_TOTAL_HINTS = ("총계", "합계")
_UNMATCHED_NOTE = "※ 미분류 신규 계정 (수동 확인 필요)"
_NUM_FORMAT = r"#,##0;\(#,##0\);\-"

_LABEL_PREFIX_RE = re.compile(
    r"^\s*(?:[ⅠⅡⅢⅣⅤⅥⅦⅧⅨⅩIVX]+\.?|\d+\.|\(\d+\)|[①②③④⑤⑥⑦⑧⑨⑩]|[가-힣]\.)\s*"
)
_PAREN_RE = re.compile(r"\(([^()]*)\)")


def normalize_label(label):
    """로마자/번호/가나다 접두어와 공백을 제거해 핵심 계정명만 남긴다."""
    if not label:
        return ""
    text = str(label).strip()
    while True:
        new_text = _LABEL_PREFIX_RE.sub("", text, count=1).strip()
        if new_text == text:
            break
        text = new_text
    return text


def _specific_keys(label):
    """괄호 안 상세 계정명에서 뽑은, 더 구체적인 매칭 키.

    "매출채권 (외상매출금)"의 "외상매출금", "보증금 (임차보증금)"의 "임차보증금"처럼
    괄호 안에 병기된 세부 항목명 — 소계/그룹 라벨과 헷갈릴 일이 적어 우선 매칭에 쓴다.
    """
    if not label:
        return set()
    keys = {normalize_label(inner) for inner in _PAREN_RE.findall(str(label))}
    keys.discard("")
    return keys


def _generic_keys(label):
    """괄호를 제거하거나 유지한 전체 라벨에서 뽑은, 상대적으로 넓은 매칭 키."""
    if not label:
        return set()
    text = str(label).strip()
    keys = {normalize_label(re.sub(r"\([^()]*\)", "", text).strip()), normalize_label(text)}
    # "부채와자본총계" ↔ "부채와 자본총계"처럼 띄어쓰기만 다른 경우도 매칭되도록 공백 제거본도 추가
    keys |= {k.replace(" ", "") for k in list(keys) if " " in k}
    keys.discard("")
    return keys


def label_match_keys(label):
    """계정명 매칭용 후보 키 전체(구체적 키 + 넓은 키)."""
    return _specific_keys(label) | _generic_keys(label)


def _is_total_label(label):
    return any(hint in (label or "") for hint in _TOTAL_HINTS)


def read_existing_raw(file_bytes):
    """기존 RAW 시트를 읽어 재무상태표/손익계산서 각각의 행 목록과 연도를 반환한다."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    if "RAW" not in wb.sheetnames:
        raise ValueError("'RAW' 시트를 찾을 수 없습니다.")
    ws = wb["RAW"]

    section = None
    year_row = None
    year_cols = {}  # col_idx -> year label(str)
    bs_rows, is_rows = [], []

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        label_cell = row[1] if len(row) > 1 else None  # column B
        label = str(label_cell.value).strip() if label_cell and label_cell.value is not None else ""

        if _SECTION_BS in label:
            section, year_row = "bs", None
            continue
        if _SECTION_IS in label:
            section, year_row = "is", None
            continue
        if section is None:
            continue

        if year_row is None:
            found = {}
            for cell in row:
                if cell.column <= 2:
                    continue
                if isinstance(cell.value, (int, float)):
                    found[cell.column] = str(int(cell.value))
                elif isinstance(cell.value, str) and re.fullmatch(r"20\d{2}", cell.value.strip()):
                    found[cell.column] = cell.value.strip()
            if found:
                year_cols = found
                year_row = row[0].row
            continue

        if not label:
            continue

        values = {}
        for col_idx, year in year_cols.items():
            cell = ws.cell(row=row[0].row, column=col_idx)
            values[year] = cell.value if isinstance(cell.value, (int, float)) else 0

        entry = {"label": label, "row_index": row[0].row, "values": values}
        (bs_rows if section == "bs" else is_rows).append(entry)

    years = list(dict.fromkeys(year_cols.values()))
    return {"years": years, "year_cols": year_cols, "bs_rows": bs_rows, "is_rows": is_rows}


def auto_match(existing_rows, new_entries):
    """자동 매칭해 검토 표에 미리 채워둘 초안 plan을 만든다.

    "매출채권 (외상매출금)" ↔ "가.외상매출금"처럼 표기가 다른 경우가 흔해서 두 단계로 매칭한다.
    1단계: 괄호 안 상세 계정명(구체적 키)끼리 먼저 매칭 — "4.매입채무"(그룹 소계) 대신
           "가.외상매입금"(세부 항목)처럼 더 정확한 대상을 우선 찾기 위함.
    2단계: 1단계에서 못 찾은 것만 넓은 키(괄호 제거/전체 라벨)로 다시 매칭.
    """
    remaining = [
        {"entry": e, "specific": _specific_keys(e.get("account", "")), "generic": _generic_keys(e.get("account", ""))}
        for e in new_entries
    ]

    plan = []
    for existing in existing_rows:
        specific = _specific_keys(existing["label"])
        generic = _generic_keys(existing["label"])
        match = None
        # 1순위: 기존 라벨의 "괄호 안 세부명"이, 후보(구체적이든 넓은 키든)와 일치하는지 먼저 본다.
        if specific:
            for i, cand in enumerate(remaining):
                if specific & (cand["specific"] | cand["generic"]):
                    match = remaining.pop(i)["entry"]
                    break
        # 2순위: "괄호 안 세부명"이 있는 계정은 넓은 키로 재시도하지 않는다 — 넓은 키로는
        # "2.단기차입금"(그룹 소계) 같은 상위 항목과 잘못 엮이기 쉽기 때문. 세부명이 있는데
        # 못 찾았으면 차라리 미매칭으로 남겨 검토 표에서 사람이 확인하게 한다.
        if match is None and not specific:
            for i, cand in enumerate(remaining):
                if generic & (cand["specific"] | cand["generic"]):
                    match = remaining.pop(i)["entry"]
                    break

        new_value = 0
        if match is not None:
            values = match.get("values", [])
            new_value = values[-1] if values else 0
        plan.append({
            "label": existing["label"],
            "row_index": existing["row_index"],
            "existing_values": existing["values"],
            "new_value": new_value,
        })

    for cand in remaining:
        entry = cand["entry"]
        values = entry.get("values", [])
        plan.append({
            "label": entry.get("account", ""),
            "row_index": None,
            "existing_values": {},
            "new_value": values[-1] if values else 0,
        })

    return plan


def preview_to_dataframe_records(plan_rows, years):
    """plan을 계정과목+연도별 컬럼의 표 형태 레코드 리스트로 변환한다 (data_editor 표시용)."""
    records = []
    for row in plan_rows:
        record = {"계정과목": row["label"]}
        for year in years[:-1]:
            record[year] = row["existing_values"].get(year, 0)
        record[years[-1]] = row["new_value"]
        records.append(record)
    return records


def plan_from_edited(existing_rows, edited_records, years):
    """사용자가 검토 표에서 수정한 결과를 최종 반영 plan으로 재구성한다 (이게 최종 소스).

    라벨이 기존 행과 일치하면 그 행을 갱신, 일치하지 않으면(사용자가 새로 추가한 행 포함)
    새 행으로 취급한다.
    """
    existing_by_label = {row["label"]: row for row in existing_rows}
    last_year = years[-1]
    plan = []
    for record in edited_records:
        label = str(record.get("계정과목", "") or "").strip()
        if not label:
            continue
        new_value = record.get(last_year, 0)
        try:
            new_value = float(new_value) if new_value not in (None, "") else 0
        except (TypeError, ValueError):
            new_value = 0
        existing_row = existing_by_label.get(label)
        if existing_row is not None:
            plan.append({
                "label": label,
                "row_index": existing_row["row_index"],
                "existing_values": existing_row["values"],
                "new_value": new_value,
                "is_new_row": False,
            })
        else:
            plan.append({
                "label": label,
                "row_index": None,
                "existing_values": {},
                "new_value": new_value,
                "is_new_row": True,
            })
    return plan


def entries_from_records(records, years):
    """(신규 파일 생성용) 표 레코드를 {"account", "values"} 리스트로 변환한다."""
    entries = []
    for record in records:
        label = str(record.get("계정과목", "") or "").strip()
        if not label:
            continue
        values = []
        for year in years:
            val = record.get(year, 0)
            try:
                val = float(val) if val not in (None, "") else 0
            except (TypeError, ValueError):
                val = 0
            values.append(val)
        entries.append({"account": label, "values": values})
    return entries


def _translate_formula(formula, old_col_letter, new_col_letter):
    """셀 참조의 열 문자만 old->new로 바꾼다 (예: C6:C9 -> D6:D9). 행 번호는 그대로 유지."""
    def repl(match):
        col, row_num = match.group(1), match.group(2)
        return f"{new_col_letter}{row_num}" if col == old_col_letter else match.group(0)
    return re.sub(r"\$?([A-Z]+)\$?(\d+)", repl, formula)


def write_merged_workbook(original_file_bytes, existing, bs_plan, is_plan, new_year):
    """기존 워크북에 새 연도 컬럼을 추가한 새 파일(bytes)을 만든다.

    매칭된 기존 행은 원래 위치에 값/서식만 채운다(행 이동 없음).
    매칭되지 않은 신규 계정(사용자가 검토 표에서 추가한 행 포함)은 각 섹션
    (재무상태표/손익계산서) 맨 끝에 "미분류 신규 계정" 표시와 함께 추가한다 —
    총계/합계 수식의 합산 범위는 자동으로 넓히지 않으므로 엑셀에서 직접 확인이 필요하다.
    """
    wb = openpyxl.load_workbook(io.BytesIO(original_file_bytes), data_only=False)
    ws = wb["RAW"]

    year_cols = existing["year_cols"]
    last_col = max(year_cols.keys())
    new_col = last_col + 1
    new_col_letter = get_column_letter(new_col)
    src_col_letter = get_column_letter(last_col)

    header_row = None
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(row=r, column=last_col).value) == str(existing["years"][-1]):
            header_row = r
            break

    # 1) 매칭된 기존 행: 값/서식 채우기 (행 이동 없음)
    for plan_rows in (bs_plan, is_plan):
        for row in plan_rows:
            if row["is_new_row"]:
                continue
            r = row["row_index"]
            old_cell = ws.cell(row=r, column=last_col)
            new_cell = ws.cell(row=r, column=new_col)
            if isinstance(old_cell.value, str) and old_cell.value.startswith("="):
                new_cell.value = _translate_formula(old_cell.value, src_col_letter, new_col_letter)
            else:
                new_cell.value = row["new_value"]
            new_cell.number_format = old_cell.number_format
            new_cell.font = copy(old_cell.font)
            new_cell.alignment = copy(old_cell.alignment)
            new_cell.border = copy(old_cell.border)

    # 2) 연도 헤더
    if header_row:
        old_header = ws.cell(row=header_row, column=last_col)
        new_header = ws.cell(row=header_row, column=new_col)
        new_header.value = int(new_year) if str(new_year).isdigit() else new_year
        new_header.font = copy(old_header.font)
        new_header.alignment = copy(old_header.alignment)
        new_header.border = copy(old_header.border)

    ws.column_dimensions[new_col_letter].width = ws.column_dimensions[src_col_letter].width

    # 3) 미매칭 신규 계정: 각 섹션 끝에 추가 (행 번호가 큰 섹션부터 삽입해야 위쪽 삽입 지점이 안 밀림)
    unmatched_notes = []
    insertions = []
    is_new = [row for row in is_plan if row["is_new_row"]]
    if existing["is_rows"] and is_new:
        insertions.append((existing["is_rows"][-1]["row_index"] + 1, is_new))
    bs_new = [row for row in bs_plan if row["is_new_row"]]
    if existing["bs_rows"] and bs_new:
        insertions.append((existing["bs_rows"][-1]["row_index"] + 1, bs_new))
    insertions.sort(key=lambda item: item[0], reverse=True)

    for insert_at, rows_to_add in insertions:
        ws.insert_rows(insert_at, amount=len(rows_to_add) + 1)
        label_font = copy(ws.cell(row=max(insert_at - 1, 1), column=2).font)
        ws.cell(row=insert_at, column=2, value=_UNMATCHED_NOTE).font = label_font
        for i, row in enumerate(rows_to_add, start=1):
            r = insert_at + i
            ws.cell(row=r, column=2, value=row["label"]).font = label_font
            cell = ws.cell(row=r, column=new_col, value=row["new_value"])
            cell.number_format = _NUM_FORMAT
            unmatched_notes.append(row["label"])

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue(), unmatched_notes


def write_new_raw_workbook(bs_entries, is_entries, years):
    """기존 RAW 파일이 없을 때, 추출 결과만으로 새 RAW 시트를 만든다."""
    from openpyxl.styles import Font, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RAW"
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 30
    for idx in range(len(years)):
        ws.column_dimensions[get_column_letter(3 + idx)].width = 16

    border = Border(bottom=Side(style="thin"))

    def write_section(start_row, title, entries):
        r = start_row
        header_cell = ws.cell(row=r, column=2, value=f"{title}           (단위: 원)")
        header_cell.font = Font(name="Noto Sans CJK SC", size=18, bold=True)
        header_cell.alignment = Alignment(horizontal="right")
        header_cell.border = border
        ws.row_dimensions[r].height = 28.5
        r += 1

        label_header = ws.cell(row=r, column=2, value="계정과목")
        label_header.font = Font(name="맑은 고딕", size=10, bold=True)
        label_header.border = border
        for idx, year in enumerate(years):
            c = ws.cell(row=r, column=3 + idx, value=int(year) if str(year).isdigit() else year)
            c.font = Font(name="맑은 고딕", size=10, bold=True)
            c.alignment = Alignment(horizontal="center")
            c.border = border
        r += 1

        for entry in entries:
            is_total = _is_total_label(entry.get("account", ""))
            label_cell = ws.cell(row=r, column=2, value=entry.get("account", ""))
            label_cell.font = Font(name="맑은 고딕", size=10, bold=is_total)
            label_cell.border = border
            values = entry.get("values", [])
            for idx in range(len(years)):
                val = values[idx] if idx < len(values) else 0
                c = ws.cell(row=r, column=3 + idx, value=val)
                c.font = Font(name="맑은 고딕", size=10, bold=False)
                c.number_format = _NUM_FORMAT
                c.alignment = Alignment(horizontal="right")
                c.border = border
            r += 1
        return r + 2

    next_row = write_section(2, _SECTION_BS, bs_entries)
    write_section(next_row, _SECTION_IS, is_entries)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
